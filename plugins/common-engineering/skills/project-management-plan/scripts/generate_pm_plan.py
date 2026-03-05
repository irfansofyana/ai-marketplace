#!/usr/bin/env python3
"""Generate a Project Management Plan Excel workbook from JSON configuration.

Usage:
    python3 generate_pm_plan.py config.json
    python3 generate_pm_plan.py config.json --output project-plan.xlsx
"""

import json
import sys
import os
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# --- Style constants ---
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="2F5496")
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
GANTT_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GANTT_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=10)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def apply_header_style(cell):
    """Apply standard header styling to a cell."""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def apply_data_style(cell, wrap=False):
    """Apply standard data styling to a cell."""
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)


def parse_date(date_str):
    """Parse a date string in YYYY-MM-DD format."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return None


def get_monday(dt):
    """Get the Monday of the week containing dt."""
    return dt - timedelta(days=dt.weekday())


def create_main_program_plan(wb, config):
    """Create the Main Program Plan tab with Gantt chart."""
    ws = wb.active
    ws.title = "Main Program Plan"

    project_name = config.get("project_name", "Project Plan")
    tasks = config.get("tasks", [])

    # --- Row 1: Project Title (merged) ---
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = project_name
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 35

    # --- Row 2: Headers ---
    headers = [
        "No", "Milestone", "Task Item", "Workstream", "PIC",
        "Status", "Start Date", "Due Date", "JIRA Link", "Remarks"
    ]
    col_widths = [5, 18, 35, 18, 15, 14, 13, 13, 20, 25]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 30

    # --- Status data validation ---
    status_dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,Blocked,On Hold"',
        allow_blank=True
    )
    status_dv.error = "Please select a valid status"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)

    # --- Populate task rows ---
    data_start_row = 3
    num_rows = max(len(tasks), 30)  # At least 30 rows for template

    for i in range(num_rows):
        row = data_start_row + i
        ws.row_dimensions[row].height = 22

        if i < len(tasks):
            task = tasks[i]
            ws.cell(row=row, column=1, value=task.get("no", i + 1))
            ws.cell(row=row, column=2, value=task.get("milestone", ""))
            ws.cell(row=row, column=3, value=task.get("task", ""))
            ws.cell(row=row, column=4, value=task.get("workstream", ""))
            ws.cell(row=row, column=5, value=task.get("pic", ""))
            ws.cell(row=row, column=6, value=task.get("status", "Not Started"))

            start_dt = parse_date(task.get("start_date", ""))
            due_dt = parse_date(task.get("due_date", ""))
            if start_dt:
                ws.cell(row=row, column=7, value=start_dt)
                ws.cell(row=row, column=7).number_format = "YYYY-MM-DD"
            if due_dt:
                ws.cell(row=row, column=8, value=due_dt)
                ws.cell(row=row, column=8).number_format = "YYYY-MM-DD"

            ws.cell(row=row, column=9, value=task.get("jira_link", ""))
            ws.cell(row=row, column=10, value=task.get("remarks", ""))
        else:
            ws.cell(row=row, column=1, value=i + 1)

        # Apply styling to all cells in row
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            apply_data_style(cell, wrap=(col in [3, 10]))
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if i % 2 == 1 and not cell.value:
                pass  # Keep alternating rows subtle

        # Apply status validation
        status_dv.add(ws.cell(row=row, column=6))

    last_data_row = data_start_row + num_rows - 1

    # --- Gantt Chart (full-year Monday-based timeline from column K onward) ---
    # Determine the Gantt year
    gantt_year = None
    project_start_date = parse_date(config.get("start_date", ""))
    if project_start_date:
        gantt_year = project_start_date.year
    else:
        task_dates = []
        for t in tasks:
            s = parse_date(t.get("start_date", ""))
            e = parse_date(t.get("due_date", ""))
            if s:
                task_dates.append(s)
            if e:
                task_dates.append(e)
        if task_dates:
            gantt_year = min(task_dates).year

    if gantt_year is None:
        gantt_year = datetime.now().year

    # Generate all Mondays from Jan 1 to Dec 31 of the gantt year
    jan1 = datetime(gantt_year, 1, 1)
    dec31 = datetime(gantt_year, 12, 31)
    first_monday = get_monday(jan1)
    if first_monday < jan1:
        first_monday += timedelta(days=7)
    weeks = []
    current = first_monday
    while current <= dec31:
        weeks.append(current)
        current += timedelta(days=7)

    gantt_start_col = 11  # Column K

    # Gantt header label in row 1
    gantt_label_cell = ws.cell(row=1, column=gantt_start_col, value="GANTT CHART")
    gantt_label_cell.font = SUBTITLE_FONT
    gantt_label_cell.alignment = Alignment(horizontal="center", vertical="center")
    if len(weeks) > 1:
        ws.merge_cells(
            start_row=1, start_column=gantt_start_col,
            end_row=1, end_column=gantt_start_col + len(weeks) - 1
        )

    # Week headers in row 2
    for w_idx, week_date in enumerate(weeks):
        col = gantt_start_col + w_idx
        cell = ws.cell(row=2, column=col, value=week_date)
        cell.number_format = "MM/DD"
        cell.font = Font(name="Calibri", bold=True, size=8, color="2F5496")
        cell.fill = GANTT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = 6.5

    # Gantt bars using conditional formatting (formula-based)
    # For each data row, each Gantt column fills if the week overlaps [start, due]
    start_col_letter = "G"  # Start Date column
    due_col_letter = "H"    # Due Date column

    for w_idx, week_date in enumerate(weeks):
        gantt_col = gantt_start_col + w_idx
        gantt_col_letter = get_column_letter(gantt_col)
        week_end = week_date + timedelta(days=6)

        for row in range(data_start_row, last_data_row + 1):
            cell = ws.cell(row=row, column=gantt_col)
            cell.border = THIN_BORDER

        # Conditional formatting: fill cell if task overlaps this week
        # Formula: AND(start_date <= week_end, due_date >= week_start)
        cell_range = f"{gantt_col_letter}{data_start_row}:{gantt_col_letter}{last_data_row}"

        # Use the week header cell reference for dynamic comparison
        week_header_ref = f"${gantt_col_letter}$2"
        # week_end = week_start + 6 days
        formula = (
            f'AND(${start_col_letter}{data_start_row}<>'
            f'"",${due_col_letter}{data_start_row}<>"",'
            f"${start_col_letter}{data_start_row}<={week_header_ref}+6,"
            f"${due_col_letter}{data_start_row}>={week_header_ref})"
        )

        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[formula], fill=GANTT_FILL)
        )

    # Freeze panes: freeze header rows and task name columns
    ws.freeze_panes = "C3"

    return ws


def create_charters(wb, config):
    """Create the Charters tab."""
    ws = wb.create_sheet("Charters")

    charters = config.get("charters", {})

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = f"{config.get('project_name', 'Project')} - Team Charter"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers
    headers = ["Core Team", "PICs"]
    col_widths = [25, 50]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 30

    # Role mappings
    roles = [
        ("Program Sponsor", charters.get("program_sponsor", "")),
        ("Business", charters.get("business", "")),
        ("Product & Design", charters.get("product_design", "")),
        ("TPM", charters.get("tpm", "")),
        ("PMO", charters.get("pmo", "")),
        ("BA", charters.get("ba", "")),
        ("Engineering Leads", charters.get("engineering_leads", "")),
        ("QA", charters.get("qa", "")),
    ]

    for i, (role, pic) in enumerate(roles):
        row = 3 + i
        ws.row_dimensions[row].height = 28

        role_cell = ws.cell(row=row, column=1, value=role)
        apply_data_style(role_cell)
        role_cell.font = Font(name="Calibri", bold=True, size=10)

        pic_cell = ws.cell(row=row, column=2, value=pic)
        apply_data_style(pic_cell, wrap=True)

        if i % 2 == 1:
            role_cell.fill = ALT_ROW_FILL
            pic_cell.fill = ALT_ROW_FILL

    ws.freeze_panes = "A3"
    return ws


def create_budget(wb, config):
    """Create the Budget tab with SUM formula."""
    ws = wb.create_sheet("Budget")

    budget_items = config.get("budget", [])

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"{config.get('project_name', 'Project')} - Budget"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers
    headers = ["Budget Category", "Budget Owner", "Amount", "Currency", "Notes"]
    col_widths = [25, 20, 18, 12, 35]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 30

    # Default items if none provided
    if not budget_items:
        budget_items = [
            {"category": "Infrastructure", "owner": "", "amount": 0, "currency": "IDR", "notes": ""},
            {"category": "Vendor", "owner": "", "amount": 0, "currency": "IDR", "notes": ""},
        ]

    # Data rows
    data_start = 3
    for i, item in enumerate(budget_items):
        row = data_start + i
        ws.row_dimensions[row].height = 24

        ws.cell(row=row, column=1, value=item.get("category", ""))
        ws.cell(row=row, column=2, value=item.get("owner", ""))

        amount_cell = ws.cell(row=row, column=3, value=item.get("amount", 0))
        amount_cell.number_format = '#,##0'

        ws.cell(row=row, column=4, value=item.get("currency", "IDR"))
        ws.cell(row=row, column=5, value=item.get("notes", ""))

        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            apply_data_style(cell, wrap=(col == 5))
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL

    last_data_row = data_start + len(budget_items) - 1

    # Total row
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 28

    total_label = ws.cell(row=total_row, column=1, value="TOTAL")
    total_label.font = TOTAL_FONT
    total_label.fill = TOTAL_FILL
    total_label.border = THIN_BORDER
    total_label.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=total_row, column=2).fill = TOTAL_FILL
    ws.cell(row=total_row, column=2).border = THIN_BORDER

    total_formula = ws.cell(
        row=total_row, column=3,
        value=f"=SUM(C{data_start}:C{last_data_row})"
    )
    total_formula.font = TOTAL_FONT
    total_formula.fill = TOTAL_FILL
    total_formula.border = THIN_BORDER
    total_formula.number_format = '#,##0'

    for col in [4, 5]:
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    ws.freeze_panes = "A3"
    return ws


def create_raid_logs(wb, config):
    """Create the RAID Logs tab."""
    ws = wb.create_sheet("RAID Logs")

    raid_items = config.get("raid", [])

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"{config.get('project_name', 'Project')} - RAID Logs"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers
    headers = ["No", "Category", "Description", "Impact", "Status", "PIC", "Action Items"]
    col_widths = [5, 14, 40, 12, 14, 15, 40]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 30

    # Data validations
    category_dv = DataValidation(
        type="list",
        formula1='"Risk,Assumption,Issue,Dependency"',
        allow_blank=True
    )
    category_dv.error = "Please select: Risk, Assumption, Issue, or Dependency"
    category_dv.errorTitle = "Invalid Category"
    ws.add_data_validation(category_dv)

    impact_dv = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    impact_dv.error = "Please select: High, Medium, or Low"
    impact_dv.errorTitle = "Invalid Impact"
    ws.add_data_validation(impact_dv)

    status_dv = DataValidation(
        type="list",
        formula1='"Open,Mitigated,Closed"',
        allow_blank=True
    )
    status_dv.error = "Please select: Open, Mitigated, or Closed"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)

    # Populate rows
    num_rows = max(len(raid_items), 20)  # At least 20 rows for template
    data_start = 3

    for i in range(num_rows):
        row = data_start + i
        ws.row_dimensions[row].height = 24

        if i < len(raid_items):
            item = raid_items[i]
            ws.cell(row=row, column=1, value=item.get("no", i + 1))
            ws.cell(row=row, column=2, value=item.get("category", ""))
            ws.cell(row=row, column=3, value=item.get("description", ""))
            ws.cell(row=row, column=4, value=item.get("impact", ""))
            ws.cell(row=row, column=5, value=item.get("status", "Open"))
            ws.cell(row=row, column=6, value=item.get("pic", ""))
            ws.cell(row=row, column=7, value=item.get("action_items", ""))
        else:
            ws.cell(row=row, column=1, value=i + 1)

        # Apply styling
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            apply_data_style(cell, wrap=(col in [3, 7]))
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply data validations
        category_dv.add(ws.cell(row=row, column=2))
        impact_dv.add(ws.cell(row=row, column=4))
        status_dv.add(ws.cell(row=row, column=5))

    ws.freeze_panes = "A3"
    return ws


def generate_workbook(config):
    """Generate the complete project management plan workbook."""
    wb = openpyxl.Workbook()

    create_main_program_plan(wb, config)
    create_charters(wb, config)
    create_budget(wb, config)
    create_raid_logs(wb, config)

    output_path = config.get("output_path", "./project-plan.xlsx")

    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 generate_pm_plan.py <config.json> [--output path.xlsx]")
        sys.exit(1)

    config_path = sys.argv[1]

    with open(config_path, "r") as f:
        config = json.load(f)

    # Override output path if provided via CLI
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            config["output_path"] = sys.argv[idx + 1]

    output = generate_workbook(config)
    print(f"Successfully generated project management plan: {output}")


if __name__ == "__main__":
    main()
